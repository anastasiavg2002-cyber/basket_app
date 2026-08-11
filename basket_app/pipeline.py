# -*- coding: utf-8 -*-

import pandas as pd
import zipfile
import shutil
import os
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


# =========================
# PROCESS SHEET
# =========================

def process_sheet(df, valid_values):

    df = df.iloc[:, [1, 2, 7, 8, 15, 16, 17]].copy()
    df.columns = list("ABCDEFG")

    mask = (df["A"] == "        ALL VALUES") & (df["B"] != 0)
    df.loc[mask, "A"] = df.loc[mask, "B"]

    df = df.drop(columns=["B"])
    df.columns = list("ABCDEF")

    df = df[["A", "D", "E", "F", "B", "C"]]

    df.insert(3, "Trips (000)", "")

    df["filter"] = df["A"].astype(str).apply(
        lambda x: 0 if x in valid_values else 1
    )

    df = df[
        df["A"].astype(str).str.lower() != "    rest"
    ]

    new_columns = [
        "",
        "FMCG trips share",
        "Trips Share",
        "Trips (000)",
        "Affinity index to FMCG",
        "target_trips_raw_CS",
        "target_trips_000_CS",
        "filter"
    ]

    df.columns = new_columns

    # Убираем ненужные строки
    df = df[df["filter"] == 1].copy()

    return df


# =========================
# MATCH PROJECT
# =========================

def match_project(name, project_name):

    name = str(name).lower()
    parts = name.strip().split()

    if len(parts) == 0:
        return False

    if parts[0] == "total" and len(parts) > 1:
        return parts[1] == project_name

    return parts[0] == project_name


# =========================
# FORMAT EXCEL
# =========================

def apply_formatting(output_file):

    wb = load_workbook(output_file)

    gray = PatternFill(
        "solid",
        fgColor="777777"
    )

    light_gray = PatternFill(
        "solid",
        fgColor="CCCCCC"
    )

    orange = PatternFill(
        "solid",
        fgColor="FFA500"
    )

    for ws in wb.worksheets:

        # =========================
        # HEADER
        # =========================

        ws.insert_rows(1)
        ws.merge_cells("B1:E1")

        cell = ws["B1"]
        cell.value = ws.title
        cell.font = Font(
            bold=True,
            size=11
        )
        cell.alignment = Alignment(
            horizontal="center"
        )

        # =========================
        # FIND COLUMNS
        # =========================

        headers = {
            ws.cell(2, col).value: col
            for col in range(
                1,
                ws.max_column + 1
            )
        }

        aff_col = headers.get(
            "Affinity index to FMCG"
        )

        trips_col = headers.get(
            "target_trips_raw_CS"
        )

        if not aff_col or not trips_col:
            continue

        # =========================
        # ROW LOOP
        # =========================

        for row in range(
            3,
            ws.max_row + 1
        ):

            try:

                aff = ws.cell(
                    row=row,
                    column=aff_col
                ).value

                trips = ws.cell(
                    row=row,
                    column=trips_col
                ).value

                # convert safely
                aff = (
                    float(aff)
                    if aff is not None
                    else None
                )

                trips = (
                    float(trips)
                    if trips is not None
                    else None
                )

                # =========================
                # ORANGE
                # =========================

                if (
                    aff is not None
                    and trips is not None
                ):

                    if (
                        aff >= 1.5
                        and trips > 30
                    ):

                        for col in range(2, 6):

                            ws.cell(
                                row=row,
                                column=col
                            ).fill = orange

                        continue

                # =========================
                # LIGHT GRAY
                # =========================

                if (
                    trips is not None
                    and trips < 15
                ):

                    for col in range(2, 6):

                        ws.cell(
                            row=row,
                            column=col
                        ).fill = light_gray

                    # delete affinity value
                    ws.cell(
                        row=row,
                        column=aff_col
                    ).value = None

                    continue

                # =========================
                # GRAY
                # =========================

                if (
                    aff is not None
                    and aff >= 1.5
                ):

                    for col in range(2, 6):

                        ws.cell(
                            row=row,
                            column=col
                        ).fill = gray

            except:

                continue

    wb.save(output_file)


# =========================
# MAIN PIPELINE
# =========================

def run_pipeline(
    projects_zip_path,
    reference_path,
    config_path,
    total_path,
    total_mapping_path,
    project_mapping_path
):

    # =========================
    # CLEAN DIRECTORIES
    # =========================

    shutil.rmtree(
        "projects",
        ignore_errors=True
    )

    shutil.rmtree(
        "results",
        ignore_errors=True
    )

    Path("projects").mkdir(
        exist_ok=True
    )

    Path("results").mkdir(
        exist_ok=True
    )

    # =========================
    # UNZIP PROJECTS
    # =========================

    with zipfile.ZipFile(
        projects_zip_path,
        "r"
    ) as z:

        z.extractall(
            "projects"
        )

    # =========================
    # SHEET CONFIG
    # =========================

    config = pd.read_excel(
        config_path
    )

    rename_map = dict(
        zip(
            config["original"],
            config["rename"]
        )
    )

    order_map = dict(
        zip(
            config["rename"],
            config["order"]
        )
    )

    # =========================
    # REFERENCE
    # =========================

    ref_df = pd.read_excel(
        reference_path,
        sheet_name="Список"
    )

    valid_values = set(
        ref_df.iloc[:, 0]
        .dropna()
        .astype(str)
    )

    # =========================
    # TOTAL
    # =========================

    total_df = pd.read_excel(
        total_path
    )

    # =========================
    # CHANNEL MAPPING
    # =========================

    mapping_df = pd.read_excel(
        total_mapping_path
    )

    mapping = dict(
        zip(
            mapping_df["original"],
            mapping_df["rename"]
        )
    )

    # =========================
    # PROJECT ↔ TOTAL MAPPING
    # =========================

    project_mapping_df = pd.read_excel(
        project_mapping_path
    )

    project_mapping = dict(
        zip(
            project_mapping_df["project_folder"],
            project_mapping_df["total_project"]
        )
    )

    # =========================
    # APPLY CHANNEL MAPPING
    # =========================

    total_df.iloc[:, 1] = (
        total_df.iloc[:, 1]
        .astype(str)
        .map(
            lambda x: mapping.get(
                x,
                x
            )
        )
    )

    # =========================
    # PATHS
    # =========================

    base_path = Path(
        "projects"
    )

    result_path = Path(
        "results"
    )

    # =========================
    # FIND PROJECT FOLDERS
    # =========================

    project_folders = []

    for path in base_path.rglob("*"):

        if (
            path.is_dir()
            and list(
                path.glob("*.xlsx")
            )
        ):

            project_folders.append(
                path
            )

    # =========================
    # PROCESS PROJECTS
    # =========================

    for project_folder in project_folders:

        sheets = []

        # =========================
        # PROJECT NAME
        # =========================

        project_folder_name = (
            project_folder.name
        )

        project_name = (
            project_mapping.get(
                project_folder_name,
                project_folder_name
            )
        )

        project_name = (
            str(project_name)
            .strip()
            .lower()
        )

        # =========================
        # PROCESS FILES
        # =========================

        for file in project_folder.glob(
            "*.xlsx"
        ):

            df = pd.read_excel(
                file
            )

            original = (
                file.stem
                .split("_")[-1]
            )

            sheet_name = (
                rename_map.get(
                    original,
                    original
                )
            )

            order = (
                order_map.get(
                    sheet_name,
                    999
                )
            )

            processed_df = process_sheet(
                df,
                valid_values
            )

            # =========================
            # TOTAL MATCH
            # =========================
            
            total_project_name = str(
                project_mapping.get(
                    project_folder_name,
                    ""
                )
            ).strip().lower()
            
            total_channel_name = str(
                sheet_name
            ).strip().lower()
            
            row = total_df[
                (
                    total_df.iloc[:, 0]
                    .astype(str)
                    .str.strip()
                    .str.lower()
                    ==
                    total_project_name
                )
                &
                (
                    total_df.iloc[:, 1]
                    .astype(str)
                    .str.strip()
                    .str.lower()
                    ==
                    total_channel_name
                )
            ]

            if not row.empty:

                trips_rp = row.iloc[0, 3]

            else:

                trips_rp = None

            # =========================
            # TRIPS (000)
            # =========================

            if (
                trips_rp is not None
                and
                "Trips Share"
                in processed_df.columns
            ):

                trips_list = [
                    trips_rp
                ]

                for val in (
                    processed_df[
                        "Trips Share"
                    ].iloc[1:]
                ):

                    try:

                        trips_list.append(
                            trips_rp
                            * float(val)
                        )

                    except:

                        trips_list.append(
                            None
                        )

                processed_df[
                    "Trips (000)"
                ] = trips_list

            # =========================
            # SAVE SHEET
            # =========================

            sheets.append(
                (
                    sheet_name,
                    order,
                    processed_df
                )
            )

        # =========================
        # EMPTY PROJECT
        # =========================

        if not sheets:
            continue

        # =========================
        # SORT SHEETS
        # =========================

        sheets = sorted(
            sheets,
            key=lambda x: x[1]
        )

        # =========================
        # OUTPUT FILE
        # =========================

        output_file = (
            result_path
            /
            f"Basket Analysis_{project_folder.name}.xlsx"
        )

        with pd.ExcelWriter(
            output_file,
            engine="openpyxl"
        ) as writer:

            for (
                sheet_name,
                _,
                df
            ) in sheets:

                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False
                )

        # =========================
        # FORMAT
        # =========================

        apply_formatting(
            output_file
        )

    # =========================
    # ZIP RESULTS
    # =========================

    output_zip = "results.zip"

    with zipfile.ZipFile(
        output_zip,
        "w",
        zipfile.ZIP_DEFLATED
    ) as z:

        for file in result_path.glob(
            "*.xlsx"
        ):

            z.write(
                file,
                arcname=file.name
            )

    return output_zip
